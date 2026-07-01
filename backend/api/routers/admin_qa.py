import logging
from typing import Any
from uuid import uuid4, uuid5, NAMESPACE_OID
import pandas as pd
import datetime
import openpyxl

def get_batches(lst: list, batch_size: int = 50):
    """Chia một list thành các list nhỏ có kích thước batch_size."""
    for i in range(0, len(lst), batch_size):
        yield lst[i:i + batch_size]

def extract_markdown(cell):
    if cell.value is None:
        return ''
        
    text_content = ''
    if hasattr(cell.value, '__iter__') and not isinstance(cell.value, str):
        for part in cell.value:
            if isinstance(part, str):
                text_content += part
            else:
                text = part.text
                font = getattr(part, 'font', None)
                if font:
                    is_bold = getattr(font, 'b', False)
                    is_italic = getattr(font, 'i', False)
                    
                    if is_bold or is_italic:
                        lines = text.split('\n')
                        formatted_lines = []
                        for line in lines:
                            if not line:
                                formatted_lines.append(line)
                                continue
                            f_line = line
                            if is_bold: f_line = f'**{f_line}**'
                            if is_italic: f_line = f'*{f_line}*'
                            formatted_lines.append(f_line)
                        text = '\n'.join(formatted_lines)
                text_content += text
    else:
        text_content = str(cell.value)
        
    lines = text_content.split('\n')
    md_lines = [line.rstrip() + '  ' if line.rstrip() else '' for line in lines]
    return '\n'.join(md_lines).strip()


from fastapi import APIRouter, Depends, HTTPException, Query, status, Request, File, UploadFile
from fastapi.responses import StreamingResponse
import asyncio
import json
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from core.event_bus import subscribe_qa_events, unsubscribe_qa_events

from db.connection import get_db
from db.models import QAStaging
from api.schemas import ErrorResponse
from core.embedder import embed, embed_batch
from core.vectordb import _get_client, setup_collection, get_collection_name
from qdrant_client import models as qmodels

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/admin",
    tags=["Admin QA"],
)

@router.get("/qa-staging/stream")
async def qa_stream(request: Request):
    """
    Endpoint SSE cho phép Frontend kết nối và nhận event realtime.
    """
    queue = await subscribe_qa_events()
    
    async def event_generator():
        try:
            while True:
                if await request.is_disconnected():
                    break
                
                # Chờ tối đa 1s để check disconnect thường xuyên
                try:
                    payload = await asyncio.wait_for(queue.get(), timeout=1.0)
                    yield f"data: {json.dumps(payload)}\n\n"
                except asyncio.TimeoutError:
                    pass
        finally:
            unsubscribe_qa_events(queue)
            
    return StreamingResponse(event_generator(), media_type="text/event-stream")

@router.get("/qa-staging", response_model=dict)
async def get_qa_staging(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    """
    Lấy danh sách các Q&A chờ duyệt.
    """
    try:
        stmt = select(QAStaging).where(QAStaging.status == "pending").order_by(QAStaging.created_at.desc()).offset(skip).limit(limit)
        result = await db.execute(stmt)
        records = result.scalars().all()
        
        # Đếm tổng
        count_stmt = select(QAStaging).where(QAStaging.status == "pending")
        count_result = await db.execute(count_stmt)
        total = len(count_result.scalars().all())

        return {
            "total": total,
            "skip": skip,
            "limit": limit,
            "data": [
                {
                    "id": r.id,
                    "question": r.question,
                    "answer": r.answer,
                    "created_at": r.created_at.isoformat()
                } for r in records
            ]
        }
    except Exception as e:
        logger.error("[AdminQA] Lỗi lấy danh sách QA Staging: %s", str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Lỗi hệ thống khi lấy danh sách chờ duyệt"
        )

@router.put("/qa-staging/{qa_id}", response_model=dict)
async def update_qa_staging(
    qa_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    """
    Cập nhật nhanh (Inline edit) câu hỏi hoặc câu trả lời.
    """
    try:
        stmt = select(QAStaging).where(QAStaging.id == qa_id)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()

        if not record:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")

        if "question" in payload:
            record.question = payload["question"]
        if "answer" in payload:
            record.answer = payload["answer"]

        await db.commit()
        return {"success": True, "message": "Cập nhật thành công"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[AdminQA] Lỗi cập nhật QA Staging %s: %s", qa_id, str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi cập nhật")

@router.delete("/qa-staging/{qa_id}", response_model=dict)
async def delete_qa_staging(
    qa_id: int,
    db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    """
    Từ chối duyệt và xóa cứng bản ghi Q&A (Zero-inbox philosophy).
    """
    try:
        stmt = select(QAStaging).where(QAStaging.id == qa_id)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()

        if not record:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")

        await db.delete(record)
        await db.commit()
        return {"success": True, "message": "Đã xóa bản ghi thành công"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[AdminQA] Lỗi xóa QA Staging %s: %s", qa_id, str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi xóa")

@router.post("/qa-staging/bulk-delete", response_model=dict)
async def bulk_delete_qa_staging(
    payload: dict,
    db: AsyncSession = Depends(get_db)
):
    ids = payload.get("ids", [])
    if not ids:
        return {"success": True}
    try:
        stmt = delete(QAStaging).where(QAStaging.id.in_(ids))
        await db.execute(stmt)
        await db.commit()
        return {"success": True, "message": f"Đã xóa {len(ids)} bản ghi"}
    except Exception as e:
        logger.error("[AdminQA] Lỗi xóa bulk QA Staging: %s", str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi xóa")

@router.post("/qa-staging/{qa_id}/approve", response_model=dict)
async def approve_qa_staging(
    qa_id: int,
    db: AsyncSession = Depends(get_db)
) -> dict[str, Any]:
    """
    Duyệt Q&A:
    1. Vector hóa câu hỏi.
    2. Lưu vào Qdrant (collection: qa_semantic_cache).
    3. Xóa cứng bản ghi khỏi staging.
    """
    try:
        stmt = select(QAStaging).where(QAStaging.id == qa_id)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()

        if not record:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")

        # 1. Vector hóa câu hỏi
        vector = embed(record.question)

        # 2. Lưu vào Qdrant
        client = _get_client()
        collection_name = "qa_semantic_cache"
        
        # Đảm bảo collection tồn tại
        setup_collection(collection_name)

        point_id = str(uuid4())
        client.upsert(
            collection_name=collection_name,
            points=[
                qmodels.PointStruct(
                    id=point_id,
                    vector={
                        "dense": vector["dense"],
                        "sparse": qmodels.SparseVector(
                            indices=vector["sparse_indices"],
                            values=vector["sparse_values"]
                        )
                    },
                    payload={
                        "question": record.question,
                        "answer": record.answer,
                        "created_at": record.created_at.isoformat(),
                        "doc_type": "qa_cache"
                    }
                )
            ]
        )
        logger.info("[AdminQA] Đã duyệt và đưa câu hỏi '%s' vào Qdrant", record.question)

        # 3. Xóa cứng bản ghi (Zero-inbox philosophy)
        await db.delete(record)
        await db.commit()

        return {"success": True, "message": "Đã duyệt và lưu vào Vector DB thành công"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[AdminQA] Lỗi khi duyệt QA Staging %s: %s", qa_id, str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi duyệt Q&A")

@router.get("/qa-approved", response_model=dict)
async def get_qa_approved():
    """Lấy danh sách các câu hỏi đã duyệt từ Qdrant."""
    try:
        client = _get_client()
        collection_name = "qa_semantic_cache"
        
        # Thử lấy data bằng scroll, nếu collection chưa tồn tại thì trả về list rỗng
        try:
            records, next_page_offset = client.scroll(
                collection_name=collection_name,
                limit=1000,
                with_payload=True
            )
        except Exception:
            return {"data": [], "total": 0}

        data = []
        for r in records:
            if r.payload and r.payload.get("doc_type") == "qa_cache":
                data.append({
                    "id": r.id,
                    "question": r.payload.get("question", ""),
                    "answer": r.payload.get("answer", ""),
                    "created_at": r.payload.get("created_at", "")
                })

        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("[AdminQA] Lỗi lấy danh sách QA Approved: %s", str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi lấy danh sách đã duyệt")


@router.put("/qa-approved/{point_id}", response_model=dict)
async def update_qa_approved(point_id: str, payload: dict):
    """Cập nhật nội dung câu hỏi/trả lời của một Q&A đã duyệt trong Qdrant."""
    try:
        client = _get_client()
        collection_name = "qa_semantic_cache"
        
        # Lấy point hiện tại
        points = client.retrieve(collection_name=collection_name, ids=[point_id], with_payload=True)
        if not points:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")
            
        point = points[0]
        current_payload = point.payload or {}
        
        new_q = payload.get("question")
        new_a = payload.get("answer")
        
        if new_q is not None and str(new_q).strip():
            current_payload["question"] = str(new_q).strip()
            new_vector = embed(str(new_q).strip())
            client.upsert(
                collection_name=collection_name,
                points=[qmodels.PointStruct(
                    id=point_id, 
                    vector={
                        "dense": new_vector["dense"],
                        "sparse": qmodels.SparseVector(
                            indices=new_vector["sparse_indices"],
                            values=new_vector["sparse_values"]
                        )
                    }, 
                    payload=current_payload
                )]
            )
        elif new_a is not None and str(new_a).strip():
            current_payload["answer"] = str(new_a).strip()
            client.set_payload(
                collection_name=collection_name,
                payload=current_payload,
                points=[point_id]
            )
            
        return {"success": True, "message": "Cập nhật thành công"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[AdminQA] Lỗi cập nhật QA Approved %s: %s", point_id, str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi cập nhật QA Approved")


@router.delete("/qa-approved/{point_id}", response_model=dict)
async def delete_qa_approved(point_id: str):
    """Xóa một câu hỏi đã duyệt khỏi Qdrant."""
    try:
        client = _get_client()
        collection_name = "qa_semantic_cache"
        client.delete(
            collection_name=collection_name,
            points_selector=[point_id]
        )
        return {"success": True, "message": "Đã xóa bản ghi thành công"}
    except Exception as e:
        logger.error("[AdminQA] Lỗi xóa QA Approved %s: %s", point_id, str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi xóa QA Approved")


@router.post("/qa-approved/bulk-delete", response_model=dict)
async def bulk_delete_qa_approved(
    payload: dict
):
    ids = payload.get("ids", [])
    if not ids:
        return {"success": True}
    try:
        client = _get_client()
        collection_name = "qa_semantic_cache"
        client.delete(
            collection_name=collection_name,
            points_selector=ids
        )
        return {"success": True, "message": f"Đã xóa {len(ids)} bản ghi"}
    except Exception as e:
        logger.error("[AdminQA] Lỗi xóa bulk QA Approved: %s", str(e))
        raise HTTPException(status_code=500, detail="Lỗi hệ thống khi xóa")


@router.post("/qa-approved/import", response_model=dict)
async def import_qa_approved(file: UploadFile = File(...)):
    """Import hàng loạt Q&A từ file Excel vào Qdrant."""
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ định dạng .xls hoặc .xlsx")

    try:
        contents = await file.read()
        import io
        
        records = []
        if file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(io.BytesIO(contents), rich_text=True)
            ws = wb.active
            
            headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value is not None]
            if 'câu hỏi' not in headers or 'trả lời' not in headers:
                raise HTTPException(status_code=400, detail="File Excel phải có ít nhất 2 cột: 'Câu hỏi' và 'Trả lời'.")
                
            q_idx = headers.index('câu hỏi')
            a_idx = headers.index('trả lời')
            
            for row in ws.iter_rows(min_row=2):
                if len(row) <= max(q_idx, a_idx):
                    continue
                q_cell = row[q_idx]
                a_cell = row[a_idx]
                if q_cell.value is None and a_cell.value is None:
                    continue
                q_text = extract_markdown(q_cell)
                a_text = extract_markdown(a_cell)
                if q_text and a_text:
                    records.append([q_text, a_text])
        else:
            df = pd.read_excel(io.BytesIO(contents))
            columns = [str(c).strip().lower() for c in df.columns]
            if 'câu hỏi' not in columns or 'trả lời' not in columns:
                raise HTTPException(
                    status_code=400, 
                    detail="File Excel phải có ít nhất 2 cột: 'Câu hỏi' và 'Trả lời'."
                )
                
            q_idx = columns.index('câu hỏi')
            a_idx = columns.index('trả lời')
            records = df.iloc[:, [q_idx, a_idx]].dropna(how="any").values.tolist()
        
        if len(records) > 1000:
            raise HTTPException(status_code=400, detail="Vượt quá giới hạn 1000 dòng mỗi lần import.")
            
        if not records:
            return {"success": True, "message": "File không có dữ liệu hợp lệ.", "count": 0}

        client = _get_client()
        collection_name = "qa_semantic_cache"
        setup_collection(collection_name)
        
        batch_size = 100
        total_upserted = 0
        
        BATCH_SIZE = 50
        for batch_records in get_batches(records, BATCH_SIZE):
            valid_records = []
            for q, a in batch_records:
                question = str(q).strip()
                answer = str(a).strip()
                if not question or not answer:
                    continue
                valid_records.append((question, answer))
                
            if not valid_records:
                continue
                
            questions = [item[0] for item in valid_records]
            # Embed nguyên một batch cùng lúc thay vì từng câu
            vectors = embed_batch(questions)
            
            batch_points = []
            for i, (question, answer) in enumerate(valid_records):
                point_id = str(uuid5(NAMESPACE_OID, question.lower()))
                vector = vectors[i]
                
                batch_points.append(
                    qmodels.PointStruct(
                        id=point_id,
                        vector={
                            "dense": vector["dense"],
                            "sparse": qmodels.SparseVector(
                                indices=vector["sparse_indices"],
                                values=vector["sparse_values"]
                            )
                        },
                        payload={
                            "question": question,
                            "answer": answer,
                            "created_at": datetime.datetime.utcnow().isoformat(),
                            "doc_type": "qa_cache"
                        }
                    )
                )
                
            client.upsert(collection_name=collection_name, points=batch_points)
            total_upserted += len(batch_points)
            logger.info(f"Đã import batch {len(batch_points)} bản ghi thành công.")
            
        return {"success": True, "message": f"Đã import thành công {total_upserted} câu hỏi vào hệ thống.", "count": total_upserted}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("[AdminQA] Lỗi import Excel: %s", str(e))
        raise HTTPException(status_code=500, detail="Lỗi khi xử lý file Excel")
